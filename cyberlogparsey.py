#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
╔══════════════════════════════════════════════════════════════════╗
║          CyberLogParser Pro 2026 — SOC Edition                  ║
║          Tek Dosya / Single File Build                          ║
║                                                                  ║
║  Kurulum:  pip install customtkinter matplotlib pandas           ║
║  Çalıştır: python CyberLogParser.py                             ║
╚══════════════════════════════════════════════════════════════════╝

Ne yapar bu program?
→ Log dosyalarını otomatik tanıyıp parse eder (Apache, Nginx, Syslog,
   SSH, Windows Event, CEF, JSON/Suricata, Cisco ASA, AWS VPC Flow vb.)
→ Tehdit tespiti yapar: Brute force, SQLi, XSS, Port Scan, DoS, vb.
→ MITRE ATT&CK tekniklerini eşleştirir
→ IOC (Indicator of Compromise) çıkartır: IP, domain, hash, CVE
→ Gerçek zamanlı log izler (tail -f gibi)
→ PDF / HTML / Excel / JSON / CSV rapor üretir
→ Matplotlib ile 6 farklı grafik çizer
"""

# ─── STANDART KÜTÜPHANE İMPORTLARI ─────────────────────────────────────────
import sys
import os
import re
import json
import csv
import gzip
import bz2
import zipfile
import math
import threading
import time
import socket
import ipaddress
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Generator
from dataclasses import dataclass, field
from enum import Enum
from collections import defaultdict, Counter
from functools import lru_cache

# ─── OTOMATİK KURULUM FONKSİYONU ────────────────────────────────────────────
def install_if_missing(packages: list):
    """
    Eksik paketleri otomatik kurar.
    Kullanıcı pip install yazmak zorunda kalmadan çalışır.
    """
    import subprocess
    for pkg in packages:
        try:
            __import__(pkg.split('>=')[0].replace('-', '_'))
        except ImportError:
            print(f"[*] Kuruluyor: {pkg}")
            try:
                subprocess.check_call(
                    [sys.executable, '-m', 'pip', 'install', pkg],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL
                )
                print(f"[+] Kuruldu: {pkg}")
            except Exception as e:
                print(f"[-] Kurulamadı {pkg}: {e}")

# Zorunlu paketleri kontrol et ve kur
print("[*] Bağımlılıklar kontrol ediliyor...")
install_if_missing([
    'customtkinter',  # Gelişmiş tkinter UI framework'ü
    'matplotlib',     # Grafik çizimi
    'Pillow',         # Resim işleme (customtkinter'ın ihtiyacı)
    'openpyxl',       # Excel (XLSX) dışa aktarma
    'reportlab',      # PDF rapor üretimi
])
print("[+] Bağımlılıklar hazır.\n")

# ─── ÜÇ TARAF KÜTÜPHANE İMPORTLARI ─────────────────────────────────────────
import customtkinter as ctk          # Modern dark-mode tkinter
from tkinter import filedialog, messagebox
import tkinter as tk

# CustomTkinter tema ayarı (dark mod + mavi aksanlar)
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# UI ölçekleme (uygulama içi fontlar/boşluklar). İstersen CMD'de ayarlayabilirsin:
# set CYBERLOGPARSER_UI_SCALE=1.30
try:
    UI_SCALE = float(os.environ.get("CYBERLOGPARSER_UI_SCALE", "1.30"))
except ValueError:
    UI_SCALE = 1.30
ctk.set_widget_scaling(UI_SCALE)
ctk.set_window_scaling(UI_SCALE)


# ═══════════════════════════════════════════════════════════════════════════════
#  BÖLÜM 1: RENK PALETİ VE TEMA
#  Tüm UI renkleri burada merkezi olarak tanımlanır.
#  Değiştirmek istersen sadece buraya bakman yeter.
# ═══════════════════════════════════════════════════════════════════════════════
COLORS = {
    'bg_main':    '#060d1a',   # En dış arka plan (en koyu)
    'bg_panel':   '#0a1628',   # Panel arka planı
    'bg_card':    '#0f1f38',   # Kart widget'ları
    'bg_input':   '#0c1930',   # Metin kutuları
    'border':     '#1a3a6a',   # Kenarlıklar
    'accent':     '#00c8ff',   # Ana vurgu rengi (cyan)
    'accent2':    '#00ffaa',   # İkincil vurgu (yeşil-cyan)
    'accent3':    '#ff6b35',   # Üçüncül vurgu (turuncu)
    'critical':   '#ff2222',   # Kritik uyarı
    'high':       '#ff8c00',   # Yüksek tehdit
    'medium':     '#ffd700',   # Orta tehdit
    'low':        '#00bfff',   # Düşük tehdit
    'info':       '#888888',   # Bilgilendirme
    'success':    '#00dd88',   # Başarı yeşili
    'text':       '#d0e4f8',   # Ana metin
    'text_dim':   '#5a7a9a',   # Soluk metin
    'text_mono':  '#00ffcc',   # Monospace metin (log çıktısı)
}

# Tehdit seviyesi → renk eşlemesi
SEVERITY_COLORS = {
    'CRITICAL': COLORS['critical'],
    'HIGH':     COLORS['high'],
    'MEDIUM':   COLORS['medium'],
    'LOW':      COLORS['low'],
    'INFO':     COLORS['info'],
}

from cyberlogparser.core.parsing import (
    BaseParser,
    FormatDetector,
    LogEntry,
    LogFileReader,
    LogFormat,
    auto_parse_file,
    get_parser,
    parse_timestamp,
)


# ═══════════════════════════════════════════════════════════════════════════════
#  BÖLÜM 8: MITRE ATT&CK FRAMEWORK
#  MITRE ATT&CK, saldırı taktik ve tekniklerinin katalogudur.
#  Her tehdit tespiti bir MITRE tekniğiyle eşleştirilir.
# ═══════════════════════════════════════════════════════════════════════════════
MITRE_TECHNIQUES = {
    'T1078':  {'name': 'Valid Accounts',                  'tactic': 'Initial Access'},
    'T1190':  {'name': 'Exploit Public-Facing App',       'tactic': 'Initial Access'},
    'T1110':  {'name': 'Brute Force',                     'tactic': 'Credential Access'},
    'T1003':  {'name': 'OS Credential Dumping',           'tactic': 'Credential Access'},
    'T1046':  {'name': 'Network Service Discovery',       'tactic': 'Discovery'},
    'T1083':  {'name': 'File and Dir Discovery',          'tactic': 'Discovery'},
    'T1053':  {'name': 'Scheduled Task/Job',              'tactic': 'Persistence'},
    'T1136':  {'name': 'Create Account',                  'tactic': 'Persistence'},
    'T1543':  {'name': 'Create/Modify System Process',    'tactic': 'Persistence'},
    'T1059':  {'name': 'Command and Scripting Interpreter','tactic': 'Execution'},
    'T1070':  {'name': 'Indicator Removal',               'tactic': 'Defense Evasion'},
    'T1562':  {'name': 'Impair Defenses',                 'tactic': 'Defense Evasion'},
    'T1498':  {'name': 'Network Denial of Service',       'tactic': 'Impact'},
    'T1048':  {'name': 'Exfiltration Over Alt Protocol',  'tactic': 'Exfiltration'},
    'T1021':  {'name': 'Remote Services',                 'tactic': 'Lateral Movement'},
}


# ═══════════════════════════════════════════════════════════════════════════════
#  BÖLÜM 9: TEHDIT TESPİT VERİ YAPILARI
# ═══════════════════════════════════════════════════════════════════════════════
class ThreatSeverity(Enum):
    """Tehdit önem seviyeleri."""
    INFO     = 0
    LOW      = 1
    MEDIUM   = 2
    HIGH     = 3
    CRITICAL = 4


@dataclass
class ThreatAlert:
    """Tespit edilen tehditin tüm bilgileri."""
    alert_id:        str = ""
    name:            str = ""
    description:     str = ""
    severity:        ThreatSeverity = ThreatSeverity.INFO
    mitre_technique: str = ""                # Örn: T1110
    mitre_tactic:    str = ""                # Örn: Credential Access
    source_ip:       str = ""
    dest_ip:         str = ""
    username:        str = ""
    hostname:        str = ""
    timestamp:       Optional[datetime] = None
    evidence:        list = field(default_factory=list)  # Delil listesi
    confidence:      float = 0.0             # Güven skoru (0-1)
    score:           float = 0.0             # Tehdit skoru (0-100)
    raw_lines:       list = field(default_factory=list)
    tags:            list = field(default_factory=list)  # Kategoriler


# ═══════════════════════════════════════════════════════════════════════════════
#  BÖLÜM 10: IOC (INDICATOR OF COMPROMISE) ÇIKARICI
#  Log satırlarından tehlike göstergelerini çıkarır.
#  IP, domain, URL, hash, CVE, email gibi bilgileri bulur.
# ═══════════════════════════════════════════════════════════════════════════════
class IOCExtractor:
    """
    Tehlike göstergesi (IOC) çıkarma sınıfı.
    Regex tabanlı pattern matching kullanır.
    """
    
    # IP adresi (IPv4)
    IP_RE       = re.compile(r'\b(?:(?:25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(?:25[0-5]|2[0-4]\d|[01]?\d\d?)\b')
    # Domain adı
    DOMAIN_RE   = re.compile(r'\b(?:[a-zA-Z0-9](?:[a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.)+(?:com|net|org|io|ru|cn|xyz|club|top|pw|cc|su|biz|info|gov|mil|edu)\b')
    # URL
    URL_RE      = re.compile(r'https?://[^\s"\'<>]+')
    # MD5 hash (32 hex karakter)
    HASH_MD5    = re.compile(r'\b[0-9a-fA-F]{32}\b')
    # SHA1 hash (40 hex karakter)
    HASH_SHA1   = re.compile(r'\b[0-9a-fA-F]{40}\b')
    # SHA256 hash (64 hex karakter)
    HASH_SHA256 = re.compile(r'\b[0-9a-fA-F]{64}\b')
    # Email adresi
    EMAIL_RE    = re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b')
    # CVE referansı (örn: CVE-2024-12345)
    CVE_RE      = re.compile(r'CVE-\d{4}-\d{4,7}')
    
    # SQL Injection kalıpları
    SQLI_PATTERNS = [
        re.compile(r"(?i)(\bunion\b.+\bselect\b|\bselect\b.+\bfrom\b|\bdrop\b.+\btable\b)"),
        re.compile(r"(?i)(1=1|1'='1|' OR '|\" OR \"|--|;--|\bOR\b\s+\d+=\d+)"),
        re.compile(r"(?i)(exec\s*\(|execute\s*\(|xp_cmdshell|sp_executesql)"),
        re.compile(r"(?i)(sleep\s*\(|benchmark\s*\(|waitfor\s+delay)"),
        re.compile(r"(?i)(information_schema|sys\.tables|sysobjects)"),
    ]
    # XSS (Cross-Site Scripting) kalıpları
    XSS_PATTERNS = [
        re.compile(r"(?i)(<script[^>]*>|</script>|javascript:|on\w+\s*=)"),
        re.compile(r"(?i)(alert\s*\(|prompt\s*\(|confirm\s*\(|eval\s*\()"),
        re.compile(r"(?i)(document\.cookie|document\.write|window\.location)"),
    ]
    # Directory Traversal kalıpları
    TRAVERSAL_PATTERNS = [
        re.compile(r"(?:\.\.[\\/]){2,}"),
        re.compile(r"(?:%2e%2e|%252e%252e)[\\/]"),
        re.compile(r"/etc/passwd|/etc/shadow|/etc/hosts"),
        re.compile(r"\\windows\\system32|\\boot\.ini"),
    ]
    # Güvenlik tarayıcı User-Agent'ları
    SCANNER_PATTERNS = [
        re.compile(r"(?i)(sqlmap|nikto|nmap|masscan|dirbuster|gobuster)"),
        re.compile(r"(?i)(w3af|openvas|nessus|burpsuite|owasp-zap)"),
        re.compile(r"(?i)(acunetix|appscan|webinspect|havij)"),
    ]
    # Zararlı yazılım kalıpları
    MALWARE_PATTERNS = [
        re.compile(r"(?i)(\.php\?cmd=|shell\.php|c99\.php|r57\.php)"),
        re.compile(r"(?i)(powershell.*-enc|-nop -exec bypass|iex\()"),
        re.compile(r"(?i)(cmd\.exe.*/c|wscript\.exe|cscript\.exe|mshta\.exe)"),
    ]
    
    # RFC 1918 özel IP adresleri (tehdit açısından daha az ilgi çekici)
    PRIVATE_NETS = [
        re.compile(r'^10\.'),
        re.compile(r'^172\.(1[6-9]|2\d|3[01])\.'),
        re.compile(r'^192\.168\.'),
        re.compile(r'^127\.'),
    ]

    @classmethod
    def is_private_ip(cls, ip: str) -> bool:
        """IP adresinin RFC 1918 özel aralığında olup olmadığını kontrol eder."""
        return any(p.match(ip) for p in cls.PRIVATE_NETS)

    @classmethod
    def extract_from_text(cls, text: str) -> dict:
        """Verilen metinden tüm IOC'leri çıkarır."""
        ips = list(set(cls.IP_RE.findall(text)))
        return {
            'ips':          ips,
            'public_ips':   [ip for ip in ips if not cls.is_private_ip(ip)],
            'domains':      list(set(cls.DOMAIN_RE.findall(text))),
            'urls':         list(set(cls.URL_RE.findall(text))),
            'hashes_md5':   list(set(cls.HASH_MD5.findall(text))),
            'hashes_sha1':  list(set(cls.HASH_SHA1.findall(text))),
            'hashes_sha256':list(set(cls.HASH_SHA256.findall(text))),
            'emails':       list(set(cls.EMAIL_RE.findall(text))),
            'cves':         list(set(cls.CVE_RE.findall(text))),
        }

    @classmethod
    def check_sqli(cls, text: str) -> bool:
        """SQL injection kalıbı var mı?"""
        return any(p.search(text) for p in cls.SQLI_PATTERNS)

    @classmethod
    def check_xss(cls, text: str) -> bool:
        """XSS kalıbı var mı?"""
        return any(p.search(text) for p in cls.XSS_PATTERNS)

    @classmethod
    def check_traversal(cls, text: str) -> bool:
        """Directory traversal kalıbı var mı?"""
        return any(p.search(text) for p in cls.TRAVERSAL_PATTERNS)

    @classmethod
    def check_scanner(cls, ua: str) -> bool:
        """Güvenlik tarayıcısı User-Agent'ı var mı?"""
        return any(p.search(ua) for p in cls.SCANNER_PATTERNS)

    @classmethod
    def check_malware(cls, text: str) -> bool:
        """Zararlı yazılım kalıbı var mı?"""
        return any(p.search(text) for p in cls.MALWARE_PATTERNS)


# ═══════════════════════════════════════════════════════════════════════════════
#  BÖLÜM 11: DAVRANIŞSAL ANALİZ (BEHAVIORAL ANALYSIS)
#  Tek satıra bakmak yetmez. Zaman içindeki davranışları takip eder.
#  Örneğin: 1 başarısız SSH girişi normal, 60 saniyede 20 tane = brute force!
# ═══════════════════════════════════════════════════════════════════════════════
class BehavioralAnalyzer:
    """
    Durum koruyan (stateful) tehdit analiz sınıfı.
    IP başına kaç başarısız giriş olduğunu, hangi portların tarandığını vb. takip eder.
    Her LogEntry için analyze_entry() çağrılır ve alert listesi döner.
    """
    
    def __init__(self):
        # {ip: [datetime, datetime, ...]} - SSH başarısız girişlerin zamanları
        self.ssh_failures: dict = defaultdict(list)
        # {ip: {username: [datetime, ...]}} - Denenen kullanıcı adları
        self.ssh_users: dict = defaultdict(lambda: defaultdict(list))
        # {ip: [datetime, ...]} - Web isteklerinin zamanları (DoS tespiti)
        self.web_requests: dict = defaultdict(list)
        # {ip: {port: True}} - Hangi portlar tarandı (port scan tespiti)
        self.dest_ports: dict = defaultdict(set)

    def analyze_entry(self, entry: LogEntry) -> list:
        """
        Tek bir log girişini analiz eder.
        Davranışsal pattern'ler için sayaçları günceller.
        Tehdit tespit edilirse ThreatAlert listesi döner.
        """
        alerts = []
        now = entry.timestamp or datetime.now()
        
        # ── SSH Brute Force Tespiti ──────────────────────────────────────────
        # Algoritma: Son 60 saniyede aynı IP'den 5+ başarısız giriş = brute force
        if entry.log_format in (LogFormat.SSH_AUTH, LogFormat.LINUX_AUTH):
            if entry.action in ('FAILED', 'INVALID_USER') and entry.source_ip:
                ip = entry.source_ip
                self.ssh_failures[ip].append(now)
                
                # 60 saniyeden eski girişleri temizle (sliding window)
                cutoff = now - timedelta(seconds=60)
                self.ssh_failures[ip] = [t for t in self.ssh_failures[ip] if t > cutoff]
                
                count = len(self.ssh_failures[ip])
                if count >= 5:  # Eşik: 5 deneme / 60 saniye
                    alerts.append(ThreatAlert(
                        alert_id='SSH-001',
                        name='SSH Brute Force Saldırısı',
                        description=f'60 saniyede {count} başarısız giriş denemesi',
                        severity=ThreatSeverity.HIGH,
                        mitre_technique='T1110',
                        mitre_tactic='Credential Access',
                        source_ip=ip,
                        timestamp=now,
                        confidence=min(0.95, 0.5 + count * 0.05),
                        score=min(100, 40 + count * 5),
                        evidence=[f'{count} SSH başarısız giriş - {ip}'],
                        tags=['brute_force', 'ssh', 'credential_attack'],
                    ))
                
                # ── SSH Password Spray Tespiti ──────────────────────────────
                # Bir IP'den çok sayıda FARKLI kullanıcı adı deneniyor = password spray
                if entry.username:
                    self.ssh_users[ip][entry.username].append(now)
                    # 5 dakika içinde 5+ farklı kullanıcı adı
                    cutoff5 = now - timedelta(seconds=300)
                    active_users = sum(
                        1 for ts_list in self.ssh_users[ip].values()
                        if any(t > cutoff5 for t in ts_list)
                    )
                    if active_users >= 5:
                        alerts.append(ThreatAlert(
                            alert_id='SSH-002',
                            name='SSH Password Spray',
                            description=f'{active_users} farklı hesaba saldırı',
                            severity=ThreatSeverity.HIGH,
                            mitre_technique='T1110',
                            mitre_tactic='Credential Access',
                            source_ip=ip,
                            timestamp=now,
                            confidence=0.85,
                            score=75,
                            evidence=[f'{active_users} farklı kullanıcı adı denendi'],
                            tags=['password_spray', 'ssh'],
                        ))
        
        # ── Web Tehditleri ───────────────────────────────────────────────────
        if entry.url or entry.method:
            ip = entry.source_ip or 'unknown'
            search_text = f"{entry.url} {entry.user_agent} {entry.message}"
            
            # DoS/Flood tespiti: 10 saniyede 200+ istek
            self.web_requests[ip].append(now)
            cutoff_dos = now - timedelta(seconds=10)
            self.web_requests[ip] = [t for t in self.web_requests[ip] if t > cutoff_dos]
            
            if len(self.web_requests[ip]) >= 200:
                alerts.append(ThreatAlert(
                    alert_id='NET-001',
                    name='DoS/Flood Saldırısı',
                    description=f'10 saniyede {len(self.web_requests[ip])} istek',
                    severity=ThreatSeverity.HIGH,
                    mitre_technique='T1498',
                    mitre_tactic='Impact',
                    source_ip=ip,
                    timestamp=now,
                    confidence=0.9,
                    score=85,
                    evidence=[f'{len(self.web_requests[ip])} req/10s'],
                    tags=['dos', 'flood'],
                ))
            
            # SQL Injection tespiti
            if IOCExtractor.check_sqli(search_text):
                alerts.append(ThreatAlert(
                    alert_id='WEB-001',
                    name='SQL Injection Girişimi',
                    description=f'{ip} adresinden SQL injection denemesi',
                    severity=ThreatSeverity.CRITICAL,
                    mitre_technique='T1190',
                    mitre_tactic='Initial Access',
                    source_ip=ip,
                    timestamp=now,
                    confidence=0.8,
                    score=90,
                    evidence=[f'URL: {(entry.url or "")[:100]}'],
                    tags=['sqli', 'web_attack'],
                ))
            
            # XSS tespiti
            if IOCExtractor.check_xss(search_text):
                alerts.append(ThreatAlert(
                    alert_id='WEB-002',
                    name='XSS Girişimi',
                    description=f'{ip} adresinden Cross-Site Scripting denemesi',
                    severity=ThreatSeverity.HIGH,
                    mitre_technique='T1190',
                    mitre_tactic='Initial Access',
                    source_ip=ip,
                    timestamp=now,
                    confidence=0.75,
                    score=80,
                    evidence=[f'URL: {(entry.url or "")[:100]}'],
                    tags=['xss', 'web_attack'],
                ))
            
            # Directory Traversal tespiti
            if IOCExtractor.check_traversal(search_text):
                alerts.append(ThreatAlert(
                    alert_id='WEB-003',
                    name='Directory Traversal',
                    description=f'{ip} adresinden path traversal denemesi',
                    severity=ThreatSeverity.HIGH,
                    mitre_technique='T1083',
                    mitre_tactic='Discovery',
                    source_ip=ip,
                    timestamp=now,
                    confidence=0.85,
                    score=75,
                    evidence=[f'URL: {(entry.url or "")[:100]}'],
                    tags=['traversal', 'web_attack'],
                ))
            
            # Güvenlik tarayıcısı tespiti
            if entry.user_agent and IOCExtractor.check_scanner(entry.user_agent):
                alerts.append(ThreatAlert(
                    alert_id='WEB-004',
                    name='Güvenlik Tarayıcısı Tespit Edildi',
                    description=f'Bilinen tarayıcı UA: {entry.user_agent[:60]}',
                    severity=ThreatSeverity.MEDIUM,
                    mitre_technique='T1046',
                    mitre_tactic='Discovery',
                    source_ip=ip,
                    timestamp=now,
                    confidence=0.95,
                    score=60,
                    evidence=[f'User-Agent: {entry.user_agent[:80]}'],
                    tags=['scanner', 'recon'],
                ))
        
        # ── Windows Event Tehditleri ─────────────────────────────────────────
        if entry.event_id:
            eid = entry.event_id
            
            if eid == 1102:  # Log silme = savunma atlatma
                alerts.append(ThreatAlert(
                    alert_id='WIN-001', name='Güvenlik Logu Silindi',
                    description='Audit log temizlendi - örtbas girişimi olabilir!',
                    severity=ThreatSeverity.CRITICAL,
                    mitre_technique='T1070', mitre_tactic='Defense Evasion',
                    hostname=entry.hostname, timestamp=now,
                    confidence=1.0, score=95,
                    evidence=['Event ID 1102: Audit Log Cleared'],
                    tags=['log_cleared', 'defense_evasion'],
                ))
            
            elif eid == 4698:  # Zamanlanmış görev = kalıcılık mekanizması
                alerts.append(ThreatAlert(
                    alert_id='WIN-002', name='Zamanlanmış Görev Oluşturuldu',
                    description='Potansiyel kalıcılık mekanizması',
                    severity=ThreatSeverity.HIGH,
                    mitre_technique='T1053', mitre_tactic='Persistence',
                    username=entry.username, hostname=entry.hostname, timestamp=now,
                    confidence=0.7, score=70,
                    evidence=['Event ID 4698: Scheduled Task Created'],
                    tags=['persistence', 'scheduled_task'],
                ))
            
            elif eid == 7045:  # Yeni servis kurulumu = kalıcılık
                alerts.append(ThreatAlert(
                    alert_id='WIN-003', name='Yeni Servis Kuruldu',
                    description='Kötü amaçlı servis kurulumu olabilir',
                    severity=ThreatSeverity.HIGH,
                    mitre_technique='T1543', mitre_tactic='Persistence',
                    hostname=entry.hostname, timestamp=now,
                    confidence=0.65, score=65,
                    evidence=['Event ID 7045: Service Installed'],
                    tags=['persistence', 'service'],
                ))
        
        # ── Zararlı Yazılım Kalıpları ────────────────────────────────────────
        if IOCExtractor.check_malware(entry.raw):
            alerts.append(ThreatAlert(
                alert_id='MAL-001', name='Potansiyel Zararlı Yazılım Aktivitesi',
                description='Şüpheli komut veya dosya yolu tespit edildi',
                severity=ThreatSeverity.CRITICAL,
                mitre_technique='T1059', mitre_tactic='Execution',
                source_ip=entry.source_ip, username=entry.username,
                timestamp=now, confidence=0.7, score=85,
                evidence=[entry.raw[:200]],
                tags=['malware', 'suspicious'],
            ))
        
        return alerts


# ═══════════════════════════════════════════════════════════════════════════════
#  BÖLÜM 12: İSTATİSTİK MOTORU
#  Parse edilen tüm entry'lerden sayısal istatistikler üretir.
#  Bu istatistikler dashboard ve grafiklerde kullanılır.
# ═══════════════════════════════════════════════════════════════════════════════
class StatisticsEngine:
    """Log entry'lerinden güvenlik istatistikleri üretir."""

    @staticmethod
    def compute(entries: list) -> dict:
        """
        Tüm entry'leri tarayarak Counter'lar ve toplamlar oluşturur.
        Döndürdüğü dict UI bileşenleri tarafından kullanılır.
        """
        stats = {
            'total_entries':    len(entries),
            'top_source_ips':   Counter(),    # IP → kaç istek
            'top_dest_ips':     Counter(),    # Hedef IP dağılımı
            'top_urls':         Counter(),    # URL → kaç istek
            'top_usernames':    Counter(),    # Kullanıcı → kaç oturum
            'status_codes':     Counter(),    # HTTP durum kodu dağılımı
            'methods':          Counter(),    # HTTP metod dağılımı
            'severity_dist':    Counter(),    # Önem seviyesi dağılımı
            'protocols':        Counter(),    # Protokol dağılımı
            'actions':          Counter(),    # Aksiyon dağılımı
            'timeline':         defaultdict(int),  # Saat → event sayısı
            'bytes_by_ip':      defaultdict(int),  # IP → toplam byte
            'errors_4xx':       0,            # HTTP 4xx hataları
            'errors_5xx':       0,            # HTTP 5xx hataları
            'success_2xx':      0,            # HTTP 2xx başarıları
            'failed_logins':    0,            # Başarısız giriş sayısı
            'successful_logins':0,            # Başarılı giriş sayısı
            'unique_ips':       set(),        # Benzersiz IP seti
            'unique_users':     set(),        # Benzersiz kullanıcı seti
        }
        
        for e in entries:
            # IP istatistikleri
            if e.source_ip:
                stats['top_source_ips'][e.source_ip] += 1
                stats['unique_ips'].add(e.source_ip)
            if e.dest_ip:
                stats['top_dest_ips'][e.dest_ip] += 1
            
            # HTTP istatistikleri
            if e.url:       stats['top_urls'][e.url] += 1
            if e.method:    stats['methods'][e.method.upper()] += 1
            if e.status_code:
                stats['status_codes'][e.status_code] += 1
                sc = e.status_code
                if   400 <= sc < 500: stats['errors_4xx'] += 1
                elif sc >= 500:       stats['errors_5xx'] += 1
                elif 200 <= sc < 300: stats['success_2xx'] += 1
            
            # Kullanıcı istatistikleri
            if e.username:
                stats['top_usernames'][e.username] += 1
                stats['unique_users'].add(e.username)
            
            # Güvenlik istatistikleri
            if e.severity: stats['severity_dist'][e.severity.lower()] += 1
            if e.protocol: stats['protocols'][e.protocol.upper()] += 1
            if e.action:   stats['actions'][e.action.upper()] += 1
            if e.bytes_sent and e.source_ip:
                stats['bytes_by_ip'][e.source_ip] += e.bytes_sent
            
            # Giriş istatistikleri
            if e.action == 'FAILED':   stats['failed_logins'] += 1
            elif e.action == 'ACCEPTED': stats['successful_logins'] += 1
            
            # Zaman serisi (saatlik)
            if e.timestamp:
                hour_key = e.timestamp.strftime('%Y-%m-%d %H:00')
                stats['timeline'][hour_key] += 1
        
        # Set'leri sayıya çevir (JSON serializable olsun)
        stats['unique_ip_count']   = len(stats['unique_ips'])
        stats['unique_user_count'] = len(stats['unique_users'])
        
        # Her kategori için Top 15 listesi oluştur
        for key in ['top_source_ips', 'top_dest_ips', 'top_urls', 'top_usernames']:
            stats[f'{key}_list'] = stats[key].most_common(15)
        
        return stats


# ═══════════════════════════════════════════════════════════════════════════════
#  BÖLÜM 13: GERÇEK ZAMANLI LOG İZLEYİCİ
#  Açık log dosyasını sürekli okur (tail -f mantığı).
#  Arka plan thread'i kullanır → UI donmaz.
# ═══════════════════════════════════════════════════════════════════════════════
class LogTailHandler:
    """
    Bir log dosyasını gerçek zamanlı olarak izler.
    Unix 'tail -f' komutuna eşdeğer Python implementasyonu.
    
    Çalışma mantığı:
    - Başlarken dosyanın sonuna atlar (geçmiş satırları okumaz)
    - Her 0.5 saniyede dosya boyutunu kontrol eder
    - Yeni satır varsa callback'i çağırır
    - Log rotation'ı da destekler (dosya küçülürse başa döner)
    """
    
    def __init__(self, filepath: str, callback, poll_interval: float = 0.5):
        self.filepath      = filepath
        self.callback      = callback      # Yeni satır için çağrılacak fonksiyon
        self.poll_interval = poll_interval  # Kontrol aralığı (saniye)
        self._stop_event   = threading.Event()
        self._thread: Optional[threading.Thread] = None
        self._position     = 0             # Dosyadaki mevcut pozisyon

    def start(self):
        """Arka plan thread'ini başlat."""
        try:
            # Mevcut dosya sonuna atla
            self._position = os.path.getsize(self.filepath)
        except OSError:
            self._position = 0
        
        self._stop_event.clear()
        self._thread = threading.Thread(target=self._tail_loop, daemon=True)
        self._thread.start()

    def stop(self):
        """İzlemeyi durdur."""
        self._stop_event.set()
        if self._thread:
            self._thread.join(timeout=2)

    def _tail_loop(self):
        """Ana izleme döngüsü - arka plan thread'inde çalışır."""
        buffer = ""  # Henüz tam olmayan satırı tutar
        
        while not self._stop_event.is_set():
            try:
                current_size = os.path.getsize(self.filepath)
                
                # Log rotation tespit: dosya küçüldü = yeni dosya başladı
                if current_size < self._position:
                    self._position = 0
                
                # Yeni içerik var mı?
                if current_size > self._position:
                    with open(self.filepath, 'r', encoding='utf-8', errors='replace') as f:
                        f.seek(self._position)
                        new_data = f.read()
                        self._position = f.tell()
                    
                    buffer += new_data
                    lines = buffer.split('\n')
                    buffer = lines[-1]  # Son satır henüz tamamlanmamış olabilir
                    
                    for line in lines[:-1]:
                        line = line.strip()
                        if line:
                            self.callback(line)
                            
            except (OSError, IOError):
                pass
            
            # Bir sonraki kontrole kadar bekle
            self._stop_event.wait(self.poll_interval)


# ═══════════════════════════════════════════════════════════════════════════════
#  BÖLÜM 14: RAPOR OLUŞTURUCU
#  Analiz sonuçlarını PDF, HTML, Excel, JSON, CSV formatlarında dışa aktarır.
# ═══════════════════════════════════════════════════════════════════════════════
class ReportGenerator:
    """
    Analiz sonuçlarını çeşitli formatlarda dışa aktaran sınıf.
    
    Desteklenen formatlar:
    - JSON: Makine okuyabilir, SIEM entegrasyonu için
    - CSV: Excel/Grafana'ya aktarmak için
    - HTML: Tarayıcıda görüntülemek için (dark mode)
    - Excel (XLSX): Yöneticiye sunmak için (renkli, çok sayfalı)
    - PDF: Resmi rapor için
    """
    
    def __init__(self, entries: list, alerts: list, stats: dict):
        self.entries = entries
        self.alerts  = alerts
        self.stats   = stats
        self.ts      = datetime.now().strftime('%Y%m%d_%H%M%S')

    def _entry_to_dict(self, e: LogEntry) -> dict:
        """LogEntry'yi JSON serileştirilebilir dict'e çevirir."""
        return {
            'line_number':  e.line_number,
            'timestamp':    e.timestamp.isoformat() if e.timestamp else None,
            'source_ip':    e.source_ip,
            'dest_ip':      e.dest_ip,
            'source_port':  e.source_port,
            'dest_port':    e.dest_port,
            'protocol':     e.protocol,
            'action':       e.action,
            'status_code':  e.status_code,
            'method':       e.method,
            'url':          e.url,
            'username':     e.username,
            'hostname':     e.hostname,
            'message':      e.message,
            'severity':     e.severity,
            'threat_score': e.threat_score,
            'tags':         e.tags,
            'log_format':   str(e.log_format.value) if e.log_format else '',
        }

    def _alert_to_dict(self, a: ThreatAlert) -> dict:
        """ThreatAlert'i dict'e çevirir."""
        return {
            'alert_id':        a.alert_id,
            'name':            a.name,
            'description':     a.description,
            'severity':        a.severity.name if hasattr(a.severity, 'name') else str(a.severity),
            'mitre_technique': a.mitre_technique,
            'mitre_tactic':    a.mitre_tactic,
            'source_ip':       a.source_ip,
            'username':        a.username,
            'timestamp':       a.timestamp.isoformat() if a.timestamp else None,
            'confidence':      a.confidence,
            'score':           a.score,
            'evidence':        a.evidence,
            'tags':            a.tags,
        }

    def export_json(self, output_path: str) -> str:
        """Tam analiz raporunu JSON olarak kaydeder."""
        report = {
            'generated_at': datetime.now().isoformat(),
            'tool':        'CyberLogParser Pro 2026',
            'summary': {
                'total_entries': self.stats.get('total_entries', 0),
                'total_alerts':  len(self.alerts),
                'unique_ips':    self.stats.get('unique_ip_count', 0),
                'unique_users':  self.stats.get('unique_user_count', 0),
            },
            'alerts':  [self._alert_to_dict(a) for a in self.alerts],
            'entries': [self._entry_to_dict(e) for e in self.entries[:5000]],
        }
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, default=str)
        return output_path

    def export_csv(self, output_path: str) -> str:
        """Log entry'lerini CSV olarak kaydeder."""
        fields = ['line_number','timestamp','log_format','source_ip','dest_ip',
                  'source_port','dest_port','protocol','action','status_code',
                  'method','url','username','hostname','message','severity','threat_score']
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
            w.writeheader()
            for e in self.entries:
                row = self._entry_to_dict(e)
                row['url'] = (row.get('url') or '')[:200]
                row['message'] = (row.get('message') or '')[:300]
                w.writerow({k: row.get(k, '') for k in fields})
        return output_path

    def export_alerts_csv(self, output_path: str) -> str:
        """Sadece alert'leri CSV olarak kaydeder."""
        fields = ['alert_id','name','severity','score','confidence',
                  'source_ip','username','timestamp','mitre_technique',
                  'mitre_tactic','description','tags']
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader()
            for a in sorted(self.alerts, key=lambda x: x.score, reverse=True):
                d = self._alert_to_dict(a)
                d['tags'] = ','.join(d.get('tags', []))
                w.writerow({k: d.get(k, '') for k in fields})
        return output_path

    def export_html(self, output_path: str) -> str:
        """Dark mode HTML raporu oluşturur."""
        n_alerts = len(self.alerts)
        total    = self.stats.get('total_entries', 0)
        unique_ips = self.stats.get('unique_ip_count', 0)
        
        sev_colors = {
            'CRITICAL': '#ff2d2d', 'HIGH': '#ff8c00',
            'MEDIUM': '#ffd700', 'LOW': '#00bfff', 'INFO': '#888'
        }
        
        # Alert satırları
        rows = ""
        for a in sorted(self.alerts, key=lambda x: x.score, reverse=True)[:50]:
            sev = a.severity.name if hasattr(a.severity, 'name') else str(a.severity)
            color = sev_colors.get(sev.upper(), '#888')
            ts = a.timestamp.strftime('%Y-%m-%d %H:%M:%S') if a.timestamp else '-'
            rows += f"""<tr>
              <td><span style="color:{color};font-weight:bold">{sev}</span></td>
              <td><code style="color:#0ff">{a.alert_id}</code></td>
              <td>{a.name}</td>
              <td>{a.source_ip or '-'}</td>
              <td><code>{a.mitre_technique}</code></td>
              <td>{ts}</td>
              <td>{a.score:.0f}</td>
            </tr>"""
        
        # Top IP satırları
        ip_rows = ""
        for ip, count in self.stats.get('top_source_ips_list', [])[:10]:
            ip_rows += f"<tr><td><code>{ip}</code></td><td>{count:,}</td></tr>"
        
        html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<title>CyberLogParser Güvenlik Raporu</title>
<style>
  * {{margin:0;padding:0;box-sizing:border-box}}
  body {{background:#060d1a;color:#c8d8f0;font-family:'Courier New',monospace;padding:20px}}
  h1 {{color:#00d4ff;font-size:2rem;margin-bottom:5px}}
  h2 {{color:#00ffcc;margin:20px 0 10px;border-bottom:1px solid #1e3a5f;padding-bottom:5px}}
  .cards {{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:15px;margin:20px 0}}
  .card {{background:#0f1b2d;border:1px solid #1e3a5f;border-radius:8px;padding:15px;text-align:center}}
  .card .num {{font-size:2.2rem;color:#00d4ff;font-weight:bold}}
  .card .lbl {{color:#5a7a9a;font-size:0.8rem;margin-top:4px}}
  table {{width:100%;border-collapse:collapse;margin-bottom:20px}}
  th {{background:#0f1b2d;color:#00ffcc;padding:8px;text-align:left;border:1px solid #1e3a5f}}
  td {{padding:7px 10px;border:1px solid #1a2d4a;font-size:0.82rem}}
  tr:nth-child(even) {{background:#080f1c}}
  tr:hover {{background:#0f1b2d}}
  code {{color:#0ff;background:#0f1b2d;padding:2px 5px;border-radius:3px}}
</style></head><body>
<h1>🛡 CyberLogParser Pro 2026 — Güvenlik Raporu</h1>
<div style="color:#5a7a9a;margin-bottom:20px">
  Oluşturma: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
</div>
<div class="cards">
  <div class="card"><div class="num">{total:,}</div><div class="lbl">Toplam Kayıt</div></div>
  <div class="card"><div class="num" style="color:#ff4444">{n_alerts}</div><div class="lbl">Tehdit Uyarısı</div></div>
  <div class="card"><div class="num">{unique_ips:,}</div><div class="lbl">Benzersiz IP</div></div>
  <div class="card"><div class="num" style="color:#fa0">{self.stats.get('errors_4xx',0):,}</div><div class="lbl">4xx Hatası</div></div>
  <div class="card"><div class="num" style="color:#f00">{self.stats.get('errors_5xx',0):,}</div><div class="lbl">5xx Hatası</div></div>
  <div class="card"><div class="num">{self.stats.get('failed_logins',0):,}</div><div class="lbl">Başarısız Giriş</div></div>
</div>
<h2>🚨 Tehdit Uyarıları</h2>
<table><tr><th>Seviye</th><th>ID</th><th>Adı</th><th>Kaynak IP</th><th>MITRE</th><th>Zaman</th><th>Skor</th></tr>
{rows}
</table>
<h2>🌍 Top Kaynak IP'ler</h2>
<table><tr><th>IP Adresi</th><th>İstek Sayısı</th></tr>
{ip_rows}
</table>
<div style="text-align:center;color:#445;margin-top:30px;font-size:0.75rem">
  CyberLogParser Pro 2026 | Advanced SOC Log Analysis Platform
</div></body></html>"""
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)
        return output_path

    def export_excel(self, output_path: str) -> str:
        """
        Çok sayfalı Excel raporu oluşturur.
        - Sayfa 1: Özet istatistikler
        - Sayfa 2: Tehdit alert'leri (renk kodlu)
        - Sayfa 3: Log kayıtları
        - Sayfa 4: Top IP'ler
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter
        except ImportError:
            # openpyxl yoksa CSV'ye düş
            return self.export_csv(output_path.replace('.xlsx', '_entries.csv'))
        
        wb = openpyxl.Workbook()
        
        # ─ Sayfa 1: Özet ─────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Özet"
        ws.append(["CyberLogParser Pro 2026 — Güvenlik Raporu"])
        ws.append(["Oluşturma Tarihi", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws.append([])
        ws.append(["Metrik", "Değer"])
        for label, val in [
            ("Toplam Log Kaydı",    self.stats.get('total_entries', 0)),
            ("Tehdit Uyarısı",      len(self.alerts)),
            ("Benzersiz Kaynak IP", self.stats.get('unique_ip_count', 0)),
            ("Benzersiz Kullanıcı", self.stats.get('unique_user_count', 0)),
            ("HTTP 4xx Hatası",     self.stats.get('errors_4xx', 0)),
            ("HTTP 5xx Hatası",     self.stats.get('errors_5xx', 0)),
            ("Başarısız Giriş",     self.stats.get('failed_logins', 0)),
            ("Başarılı Giriş",      self.stats.get('successful_logins', 0)),
        ]:
            ws.append([label, val])
        
        # ─ Sayfa 2: Alert'ler ────────────────────────────────────────────
        ws2 = wb.create_sheet("Tehdit Uyarıları")
        ws2.sheet_properties.tabColor = "FF4444"
        
        headers = ['Alert ID','Ad','Seviye','Skor','Güven','Kaynak IP',
                   'Kullanıcı','MITRE','Taktik','Zaman','Açıklama']
        ws2.append(headers)
        
        # Severity renkleri (Excel cell fill)
        fills = {
            'CRITICAL': PatternFill("solid", fgColor="8B0000"),
            'HIGH':     PatternFill("solid", fgColor="8B4500"),
            'MEDIUM':   PatternFill("solid", fgColor="8B8B00"),
            'LOW':      PatternFill("solid", fgColor="004488"),
            'INFO':     PatternFill("solid", fgColor="333333"),
        }
        white_font = Font(color="FFFFFF", name='Consolas', size=9)
        
        for a in sorted(self.alerts, key=lambda x: x.score, reverse=True):
            sev = a.severity.name if hasattr(a.severity, 'name') else str(a.severity)
            row_data = [
                a.alert_id, a.name, sev, round(a.score,1), round(a.confidence,2),
                a.source_ip, a.username, a.mitre_technique, a.mitre_tactic,
                a.timestamp.strftime('%Y-%m-%d %H:%M:%S') if a.timestamp else '',
                a.description,
            ]
            ws2.append(row_data)
            # Satırı renklendir
            fill = fills.get(sev.upper(), fills['INFO'])
            for cell in ws2[ws2.max_row]:
                cell.fill = fill
                cell.font = white_font
        
        # Sütun genişlikleri otomatik ayarla
        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 50)
        
        # ─ Sayfa 3: Log Kayıtları ────────────────────────────────────────
        ws3 = wb.create_sheet("Log Kayıtları")
        ws3.append(['#','Zaman','Format','Kaynak IP','Hedef IP',
                    'S.Port','H.Port','Protokol','Aksiyon','HTTP Status',
                    'Metod','URL','Kullanıcı','Host','Mesaj'])
        
        for e in self.entries[:10000]:  # İlk 10000 satır
            ws3.append([
                e.line_number,
                e.timestamp.strftime('%Y-%m-%d %H:%M:%S') if e.timestamp else '',
                str(e.log_format.value) if e.log_format else '',
                e.source_ip, e.dest_ip, e.source_port, e.dest_port,
                e.protocol, e.action, e.status_code, e.method,
                (e.url or '')[:150], e.username, e.hostname,
                (e.message or '')[:200],
            ])
        
        # ─ Sayfa 4: Top IP'ler ───────────────────────────────────────────
        ws4 = wb.create_sheet("Top IP'ler")
        ws4.append(['Sıra','IP Adresi','İstek Sayısı'])
        for rank, (ip, count) in enumerate(self.stats.get('top_source_ips_list', []), 1):
            ws4.append([rank, ip, count])
        
        wb.save(output_path)
        return output_path

    def export_pdf(self, output_path: str) -> str:
        """
        PDF raporu oluşturur.
        reportlab kütüphanesi gereklidir.
        Yoksa HTML'e düşer.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        except ImportError:
            # reportlab yoksa HTML rapor üret
            html_path = output_path.replace('.pdf', '.html')
            return self.export_html(html_path)
        
        doc = SimpleDocTemplate(output_path, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        story  = []
        
        # Başlık
        title_style = ParagraphStyle('t', parent=styles['Title'],
                                    textColor=colors.HexColor('#00D4FF'), fontSize=20)
        story.append(Paragraph("🛡 CyberLogParser Pro 2026", title_style))
        story.append(Paragraph("Güvenlik Analiz Raporu", styles['Heading2']))
        story.append(Paragraph(f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                              styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Özet tablo
        summary = [
            ['Metrik', 'Değer'],
            ['Toplam Kayıt',     f"{self.stats.get('total_entries', 0):,}"],
            ['Tehdit Uyarısı',   str(len(self.alerts))],
            ['Benzersiz IP',     str(self.stats.get('unique_ip_count', 0))],
            ['4xx Hataları',     str(self.stats.get('errors_4xx', 0))],
            ['5xx Hataları',     str(self.stats.get('errors_5xx', 0))],
        ]
        t = Table(summary, colWidths=[8*cm, 6*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F1B2D')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#00FFCC')),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#1E3A5F')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1),
             [colors.HexColor('#0A0E1A'), colors.HexColor('#080F1C')]),
            ('TEXTCOLOR', (0,1), (-1,-1), colors.HexColor('#C8D8F0')),
        ]))
        story.append(t)
        
        doc.build(story)
        return output_path


# ═══════════════════════════════════════════════════════════════════════════════
#  BÖLÜM 15: UI BİLEŞENLERİ
#  CustomTkinter tabanlı arayüz widget'ları.
#  Her sınıf tek bir UI bileşenini temsil eder.
# ═══════════════════════════════════════════════════════════════════════════════

class StatCard(ctk.CTkFrame):
    """
    Dashboard'daki metrik kartı widget'ı.
    Büyük sayı + başlık + renkli ikon gösterir.
    
    Kullanım:
        card = StatCard(parent, "Toplam Entry", "1,234", color="#00c8ff", icon="📄")
        card.update_value("5,678")  # Değeri güncelle
    """
    def __init__(self, parent, title: str, value: str = "0",
                 color: str = "#00c8ff", icon: str = "●", **kwargs):
        super().__init__(parent, fg_color=COLORS['bg_card'],
                        corner_radius=10, border_width=1,
                        border_color=COLORS['border'], **kwargs)
        self.grid_columnconfigure(0, weight=1)
        
        # İkon
        ctk.CTkLabel(self, text=icon, font=("Consolas", 22),
                    text_color=color).grid(row=0, column=0, pady=(12,0))
        # Sayı değeri
        self._val = ctk.CTkLabel(self, text=value,
                                font=("Consolas", 28, "bold"),
                                text_color=color)
        self._val.grid(row=1, column=0)
        # Başlık
        ctk.CTkLabel(self, text=title, font=("Consolas", 10),
                    text_color=COLORS['text_dim']).grid(row=2, column=0, pady=(0,12))

    def update_value(self, value: str):
        """Kart değerini günceller (thread-safe değil, main thread'den çağır)."""
        self._val.configure(text=str(value))


class AlertsPanel(ctk.CTkScrollableFrame):
    """
    Tehdit alert'lerini listeleyen kaydırılabilir panel.
    Her alert renk kodlu kart olarak gösterilir.
    Kritik = kırmızı, High = turuncu, Medium = sarı, Low = mavi
    """
    def __init__(self, parent, **kwargs):
        super().__init__(parent, fg_color=COLORS['bg_panel'],
                        scrollbar_button_color=COLORS['border'],
                        scrollbar_button_hover_color=COLORS['accent'], **kwargs)
        self.grid_columnconfigure(0, weight=1)

    def populate(self, alerts: list):
        """Alert listesini temizle ve yeniden doldur."""
        # Mevcut widget'ları temizle
        for widget in self.winfo_children():
            widget.destroy()
        
        if not alerts:
            ctk.CTkLabel(self, text="✅  Tehdit tespit edilmedi",
                        font=("Consolas", 14),
                        text_color=COLORS['success']).grid(row=0, column=0, pady=40)
            return
        
        # Skora göre sırala (en tehditkâr üstte)
        sorted_alerts = sorted(alerts, key=lambda a: a.score, reverse=True)
        
        for i, alert in enumerate(sorted_alerts):
            sev_name = alert.severity.name if hasattr(alert.severity, 'name') else str(alert.severity)
            sev_color = SEVERITY_COLORS.get(sev_name.upper(), COLORS['info'])
            
            # Ana kart
            card = ctk.CTkFrame(self, fg_color=COLORS['bg_card'],
                               corner_radius=8, border_width=1,
                               border_color=sev_color)
            card.grid(row=i, column=0, sticky='ew', pady=3, padx=2)
            card.grid_columnconfigure(1, weight=1)
            
            # Sol renkli şerit (severity göstergesi)
            ctk.CTkFrame(card, width=4, fg_color=sev_color,
                        corner_radius=2).grid(row=0, column=0, rowspan=4,
                        sticky='ns', padx=(6,0), pady=4)
            
            # Başlık satırı
            hf = ctk.CTkFrame(card, fg_color='transparent')
            hf.grid(row=0, column=1, sticky='ew', padx=8, pady=(6,0))
            hf.grid_columnconfigure(1, weight=1)
            
            # Alert ID
            ctk.CTkLabel(hf, text=alert.alert_id,
                         font=("Consolas", 10, "bold"),
                         text_color=COLORS['accent']).grid(row=0, column=0)
            
            # Alert adı
            ctk.CTkLabel(hf, text=f"  {alert.name}",
                         font=("Consolas", 11, "bold"),
                         text_color=COLORS['text']).grid(row=0, column=1, sticky='w')
            
            # Severity badge
            ctk.CTkLabel(hf, text=f" {sev_name} ",
                         font=("Consolas", 9, "bold"),
                         text_color='#000' if sev_name == 'MEDIUM' else '#fff',
                         fg_color=sev_color, corner_radius=4).grid(row=0, column=2, padx=4)
            
            # Skor göstergesi
            ctk.CTkLabel(hf, text=f"Skor: {alert.score:.0f}/100",
                         font=("Consolas", 9),
                         text_color=sev_color).grid(row=0, column=3, padx=8)
            
            # Detay satırı (IP, kullanıcı, MITRE, zaman)
            info_parts = []
            if alert.source_ip:       info_parts.append(f"🌍 {alert.source_ip}")
            if alert.username:         info_parts.append(f"👤 {alert.username}")
            if alert.mitre_technique:  info_parts.append(f"⚔ {alert.mitre_technique}")
            if alert.mitre_tactic:     info_parts.append(f"📋 {alert.mitre_tactic}")
            if alert.timestamp:
                info_parts.append(f"🕐 {alert.timestamp.strftime('%Y-%m-%d %H:%M:%S')}")
            
            ctk.CTkLabel(card, text="  •  ".join(info_parts),
                         font=("Consolas", 9),
                         text_color=COLORS['text_dim']).grid(
                         row=1, column=1, sticky='w', padx=8)
            
            # Açıklama
            ctk.CTkLabel(card, text=f"  {alert.description}",
                         font=("Consolas", 9), text_color=COLORS['text'],
                         wraplength=700, justify='left').grid(
                         row=2, column=1, sticky='w', padx=8, pady=(0,2))
            
            # Deliller
            for j, ev in enumerate(alert.evidence[:2]):
                ctk.CTkLabel(card, text=f"  → {ev[:120]}",
                            font=("Consolas", 8),
                            text_color=COLORS['text_dim']).grid(
                            row=3+j, column=1, sticky='w', padx=8)
            
            # Etiketler
            if alert.tags:
                tf = ctk.CTkFrame(card, fg_color='transparent')
                tf.grid(row=5, column=1, sticky='w', padx=8, pady=(2,6))
                for tag in alert.tags[:6]:
                    ctk.CTkLabel(tf, text=f" {tag} ",
                                font=("Consolas", 8),
                                fg_color=COLORS['bg_panel'],
                                text_color=COLORS['accent2'],
                                corner_radius=3).pack(side='left', padx=2)


class DashboardPanel(ctk.CTkFrame):
    """
    Ana dashboard paneli.
    6 metrik kartı + Top IP listesi + Top URL listesi içerir.
    """
    def __init__(self, parent, **kwargs):
        super().__init__(parent, fg_color=COLORS['bg_panel'],
                        corner_radius=10, **kwargs)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._build()

    def _build(self):
        # ─ Üst: 6 metrik kartı ────────────────────────────────────────────
        cards_row = ctk.CTkFrame(self, fg_color='transparent')
        cards_row.grid(row=0, column=0, sticky='ew', padx=10, pady=10)
        
        # 6 kart tanımla (başlık, renk, ikon)
        card_defs = [
            ("Toplam Kayıt",     COLORS['accent'],   "📄"),
            ("Tehdit Uyarısı",   COLORS['critical'], "🚨"),
            ("Benzersiz IP",     COLORS['accent2'],  "🌍"),
            ("Kullanıcı",        COLORS['medium'],   "👤"),
            ("4xx Hata",         COLORS['high'],     "⚠"),
            ("5xx Hata",         COLORS['critical'], "❌"),
        ]
        
        self.cards = []
        for i, (title, color, icon) in enumerate(card_defs):
            c = StatCard(cards_row, title, color=color, icon=icon)
            c.grid(row=0, column=i, padx=5, sticky='ew')
            cards_row.grid_columnconfigure(i, weight=1)
            self.cards.append(c)
        
        # ─ Alt: İki tablo yan yana ────────────────────────────────────────
        bottom = ctk.CTkFrame(self, fg_color='transparent')
        bottom.grid(row=1, column=0, sticky='nsew', padx=10, pady=(0,10))
        bottom.grid_columnconfigure(0, weight=1)
        bottom.grid_columnconfigure(1, weight=1)
        bottom.grid_rowconfigure(0, weight=1)
        
        # Top IP tablosu
        self.ip_frame = self._make_table(bottom, "🌍 Top Kaynak IP'ler")
        self.ip_frame.grid(row=0, column=0, sticky='nsew', padx=(0,5))
        
        # Top URL tablosu
        self.url_frame = self._make_table(bottom, "🔗 Top URL'ler")
        self.url_frame.grid(row=0, column=1, sticky='nsew')

    def _make_table(self, parent, title: str) -> ctk.CTkFrame:
        """Başlıklı kaydırılabilir tablo frame'i oluşturur."""
        frame = ctk.CTkFrame(parent, fg_color=COLORS['bg_card'],
                            corner_radius=8, border_width=1,
                            border_color=COLORS['border'])
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(1, weight=1)
        
        ctk.CTkLabel(frame, text=title, font=("Consolas", 11, "bold"),
                    text_color=COLORS['accent']).grid(row=0, column=0,
                    sticky='w', padx=10, pady=6)
        
        scroll = ctk.CTkScrollableFrame(frame, fg_color='transparent',
                                       scrollbar_button_color=COLORS['border'])
        scroll.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        scroll.grid_columnconfigure(0, weight=1)
        frame._table = scroll  # İç referans
        return frame

    def update(self, stats: dict, alerts: list):
        """Dashboard verilerini günceller. Her parse sonrasında çağrılır."""
        # Kartları güncelle
        values = [
            f"{stats.get('total_entries', 0):,}",
            str(len(alerts)),
            str(stats.get('unique_ip_count', 0)),
            str(stats.get('unique_user_count', 0)),
            f"{stats.get('errors_4xx', 0):,}",
            f"{stats.get('errors_5xx', 0):,}",
        ]
        for card, val in zip(self.cards, values):
            card.update_value(val)
        
        # Top IP tablosunu güncelle
        for w in self.ip_frame._table.winfo_children():
            w.destroy()
        for rank, (ip, count) in enumerate(stats.get('top_source_ips_list', [])[:15]):
            row = ctk.CTkFrame(self.ip_frame._table, fg_color='transparent')
            row.grid(row=rank, column=0, sticky='ew', pady=1)
            ctk.CTkLabel(row, text=f"{rank+1:2d}.", width=25,
                        font=("Consolas", 9), text_color=COLORS['text_dim']).pack(side='left')
            ctk.CTkLabel(row, text=ip, width=140,
                        font=("Consolas", 9), text_color=COLORS['accent']).pack(side='left')
            ctk.CTkLabel(row, text=f"{count:,}",
                        font=("Consolas", 9, "bold"),
                        text_color=COLORS['accent2']).pack(side='left', padx=5)
        
        # Top URL tablosunu güncelle
        for w in self.url_frame._table.winfo_children():
            w.destroy()
        for rank, (url, count) in enumerate(stats.get('top_urls_list', [])[:15]):
            row = ctk.CTkFrame(self.url_frame._table, fg_color='transparent')
            row.grid(row=rank, column=0, sticky='ew', pady=1)
            ctk.CTkLabel(row, text=f"{rank+1:2d}.", width=25,
                        font=("Consolas", 9), text_color=COLORS['text_dim']).pack(side='left')
            short = (url or '')[:55] + ('...' if len(url or '') > 55 else '')
            ctk.CTkLabel(row, text=short, width=280,
                        font=("Consolas", 9), text_color=COLORS['text'],
                        anchor='w').pack(side='left')
            ctk.CTkLabel(row, text=str(count),
                        font=("Consolas", 9, "bold"),
                        text_color=COLORS['accent2']).pack(side='left', padx=5)


class LogViewPanel(ctk.CTkFrame):
    """
    Log entry'lerini tablo formatında gösteren kaydırılabilir panel.
    Filtreleme desteği var: IP, URL, kullanıcı, mesaj içinde arama yapılabilir.
    """
    def __init__(self, parent, **kwargs):
        super().__init__(parent, fg_color=COLORS['bg_panel'], **kwargs)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._all_entries  = []  # Tüm entry'ler (filter için saklanır)
        self._build()

    def _build(self):
        # ─ Filtre çubuğu ──────────────────────────────────────────────────
        filter_bar = ctk.CTkFrame(self, fg_color=COLORS['bg_card'],
                                 corner_radius=8)
        filter_bar.grid(row=0, column=0, sticky='ew', pady=(0,5))
        filter_bar.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(filter_bar, text="🔎 Filtre:",
                    font=("Consolas", 10, "bold"),
                    text_color=COLORS['accent']).grid(row=0, column=0, padx=10, pady=6)
        
        self._search_entry = ctk.CTkEntry(filter_bar,
                                         placeholder_text="IP, URL, kullanıcı, mesaj ara...",
                                         font=("Consolas", 10),
                                         fg_color=COLORS['bg_input'],
                                         border_color=COLORS['border'],
                                         text_color=COLORS['text'])
        self._search_entry.grid(row=0, column=1, padx=5, pady=6, sticky='ew')
        self._search_entry.bind('<Return>', lambda e: self._apply_filter())
        
        # Action filtresi
        self._action_var = ctk.StringVar(value="TÜMÜ")
        ctk.CTkOptionMenu(filter_bar,
                         values=["TÜMÜ","ACCEPTED","FAILED","INVALID_USER",
                                "BLOCKED","ALLOWED","DENY"],
                         variable=self._action_var,
                         font=("Consolas", 10),
                         fg_color=COLORS['bg_input'],
                         button_color=COLORS['border'],
                         dropdown_fg_color=COLORS['bg_card'],
                         command=lambda _: self._apply_filter()).grid(
                         row=0, column=2, padx=5)
        
        ctk.CTkButton(filter_bar, text="Ara",
                     font=("Consolas", 10),
                     fg_color=COLORS['accent'], text_color='#000',
                     command=self._apply_filter).grid(row=0, column=3, padx=5)
        
        ctk.CTkButton(filter_bar, text="Temizle",
                     font=("Consolas", 10),
                     fg_color=COLORS['bg_input'],
                     border_color=COLORS['border'], border_width=1,
                     hover_color=COLORS['border'],
                     command=self._clear_filter).grid(row=0, column=4, padx=5)
        
        self._count_lbl = ctk.CTkLabel(filter_bar, text="",
                                       font=("Consolas", 9),
                                       text_color=COLORS['text_dim'])
        self._count_lbl.grid(row=0, column=5, padx=10)
        
        # ─ Log tablosu ────────────────────────────────────────────────────
        self._scroll = ctk.CTkScrollableFrame(self, fg_color=COLORS['bg_panel'],
                                             scrollbar_button_color=COLORS['border'])
        self._scroll.grid(row=1, column=0, sticky='nsew')
        self._scroll.grid_columnconfigure(0, weight=1)

    def _apply_filter(self):
        """Arama kutusuna göre entry'leri filtrele ve yeniden göster."""
        search = self._search_entry.get().lower()
        action = self._action_var.get()
        
        filtered = self._all_entries
        
        if search:
            filtered = [e for e in filtered if
                       search in (e.source_ip or '').lower() or
                       search in (e.dest_ip or '').lower() or
                       search in (e.url or '').lower() or
                       search in (e.username or '').lower() or
                       search in (e.message or '').lower() or
                       search in e.raw.lower()]
        
        if action != "TÜMÜ":
            filtered = [e for e in filtered
                       if (e.action or '').upper() == action]
        
        self._render(filtered)
        self._count_lbl.configure(
            text=f"{len(filtered):,} / {len(self._all_entries):,} gösteriliyor")

    def _clear_filter(self):
        """Filtreyi temizle, tüm entry'leri göster."""
        self._search_entry.delete(0, 'end')
        self._action_var.set("TÜMÜ")
        self._render(self._all_entries)
        self._count_lbl.configure(text="")

    def populate(self, entries: list):
        """Entry listesini kaydet ve göster."""
        self._all_entries = entries
        self._render(entries)
        self._count_lbl.configure(text=f"{len(entries):,} kayıt")

    def _render(self, entries: list, max_rows: int = 1500):
        """Entry'leri tablo olarak çizer. max_rows performans sınırı."""
        # Eski satırları temizle
        for w in self._scroll.winfo_children():
            w.destroy()
        
        if not entries:
            ctk.CTkLabel(self._scroll, text="📭 Gösterilecek kayıt yok",
                        font=("Consolas", 12),
                        text_color=COLORS['text_dim']).grid(row=0, column=0, pady=30)
            return
        
        # Sütun tanımları (isim, genişlik)
        cols = [
            ('#', 50),
            ('Zaman', 155),
            ('Kaynak IP', 120),
            ('Hedef IP', 120),
            ('H.Port', 70),
            ('Aksiyon', 100),
            ('Status', 65),
            ('Metod', 65),
            ('Kullanıcı', 100),
            ('Mesaj', 320),
        ]
        
        # Başlık satırı
        hdr = ctk.CTkFrame(self._scroll, fg_color=COLORS['bg_card'], corner_radius=6)
        hdr.grid(row=0, column=0, sticky='ew', pady=(0,2))
        for j, (col, w) in enumerate(cols):
            ctk.CTkLabel(hdr, text=col, font=("Consolas", 10, "bold"),
                        text_color=COLORS['accent2'], width=w, anchor='w').grid(
                        row=0, column=j, padx=4, pady=4)
        
        # Veri satırları (alternatif arka plan rengi ile)
        for i, e in enumerate(entries[:max_rows]):
            ts  = e.timestamp.strftime('%m-%d %H:%M:%S') if e.timestamp else '-'
            msg = (e.message or e.url or e.raw or '')[:70]
            row_bg = COLORS['bg_panel'] if i % 2 == 0 else COLORS['bg_card']
            
            row_frame = ctk.CTkFrame(self._scroll, fg_color=row_bg, corner_radius=0)
            row_frame.grid(row=i+1, column=0, sticky='ew')
            
            values = [
                str(e.line_number),
                ts,
                e.source_ip or '-',
                e.dest_ip or '-',
                str(e.dest_port) if e.dest_port else '-',
                e.action or '-',
                str(e.status_code) if e.status_code else '-',
                e.method or '-',
                e.username or '-',
                msg,
            ]
            
            for j, (val, (_, w)) in enumerate(zip(values, cols)):
                # Değere göre renk belirle
                color = COLORS['text']
                if j == 0:    color = COLORS['text_dim']  # Satır numarası
                elif j == 5:  # Aksiyon rengi
                    if   val == 'FAILED':       color = COLORS['high']
                    elif val == 'ACCEPTED':     color = COLORS['success']
                    elif val == 'INVALID_USER': color = COLORS['high']
                    elif val == 'BLOCKED':      color = COLORS['medium']
                elif j == 6:  # HTTP status kodu rengi
                    try:
                        sc = int(val)
                        if   sc >= 500: color = COLORS['critical']
                        elif sc >= 400: color = COLORS['high']
                        elif sc >= 200: color = COLORS['success']
                    except: pass
                
                ctk.CTkLabel(row_frame, text=val, width=w,
                            font=("Consolas", 9), text_color=color,
                            anchor='w').grid(row=0, column=j, padx=4, pady=2)
        
        # Limit uyarısı
        if len(entries) > max_rows:
            ctk.CTkLabel(self._scroll,
                        text=f"⚠ İlk {max_rows:,} kayıt gösteriliyor (toplam: {len(entries):,})",
                        font=("Consolas", 9), text_color=COLORS['high']
                        ).grid(row=max_rows+1, column=0, pady=5)


class IOCPanel(ctk.CTkFrame):
    """
    IOC (Indicator of Compromise) çıkarım paneli.
    Tüm log'lardan IP, domain, URL, hash, CVE, email çıkarır.
    SQL injection, XSS, traversal, scanner ve malware pattern kontrolü yapar.
    """
    def __init__(self, parent, **kwargs):
        super().__init__(parent, fg_color=COLORS['bg_panel'], **kwargs)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        ctk.CTkLabel(self, text="🔍 IOC ÇIKARICI & TEHDİT İSTİHBARATI",
                    font=("Consolas", 12, "bold"),
                    text_color=COLORS['accent']).grid(
                    row=0, column=0, sticky='w', padx=15, pady=10)
        
        # Monospace metin kutusu (kopyalanabilir IOC listesi)
        self._text = ctk.CTkTextbox(self, font=("Consolas", 9),
                                   fg_color=COLORS['bg_input'],
                                   text_color=COLORS['text_mono'],
                                   wrap='word')
        self._text.grid(row=1, column=0, sticky='nsew', padx=10, pady=(0,10))

    def populate(self, entries: list):
        """Entry'lerden IOC'leri çıkar ve göster."""
        # Tüm ham satırları birleştir (max 5000 entry)
        all_text = " ".join(e.raw for e in entries[:5000])
        iocs = IOCExtractor.extract_from_text(all_text)
        
        self._text.configure(state='normal')
        self._text.delete('1.0', 'end')
        
        # Her IOC kategorisi için başlık + liste
        sections = [
            ("🌐 PUBLIC IP ADRESLERİ",  'public_ips',    COLORS['accent']),
            ("🏠 TÜM IP ADRESLERİ",     'ips',           COLORS['text_dim']),
            ("🔗 DOMAIN ADLARI",         'domains',       COLORS['accent2']),
            ("🌍 URL'LER",               'urls',          COLORS['medium']),
            ("🔑 MD5 HASH'LER",          'hashes_md5',    COLORS['high']),
            ("🔑 SHA1 HASH'LER",         'hashes_sha1',   COLORS['high']),
            ("🔑 SHA256 HASH'LER",       'hashes_sha256', COLORS['high']),
            ("📧 EMAIL ADRESLERİ",        'emails',        COLORS['accent']),
            ("🐛 CVE REFERANSLARI",       'cves',          COLORS['critical']),
        ]
        
        for title, key, _ in sections:
            items = iocs.get(key, [])
            if items:
                self._text.insert('end', f"\n{'━'*60}\n{title} ({len(items)})\n{'━'*60}\n")
                for item in sorted(set(items))[:100]:
                    self._text.insert('end', f"  {item}\n")
        
        # Saldırı pattern analizi
        self._text.insert('end', f"\n{'━'*60}\n🚨 SALDIRI PATTERN ANALİZİ\n{'━'*60}\n")
        
        # Her entry'yi kontrol et
        sqli  = sum(1 for e in entries if IOCExtractor.check_sqli(e.raw))
        xss   = sum(1 for e in entries if IOCExtractor.check_xss(e.raw))
        trav  = sum(1 for e in entries if IOCExtractor.check_traversal(e.raw))
        scan  = sum(1 for e in entries if IOCExtractor.check_scanner(e.user_agent or ''))
        malw  = sum(1 for e in entries if IOCExtractor.check_malware(e.raw))
        
        for name, count in [
            ("SQL Injection girişimi",  sqli),
            ("XSS girişimi",           xss),
            ("Directory traversal",    trav),
            ("Güvenlik tarayıcısı",    scan),
            ("Zararlı yazılım izi",    malw),
        ]:
            icon = "🔴" if count > 0 else "🟢"
            self._text.insert('end', f"  {icon} {name}: {count} tespit\n")
        
        self._text.configure(state='disabled')


class LiveMonitorPanel(ctk.CTkFrame):
    """
    Gerçek zamanlı log izleme paneli.
    Seçilen dosyayı tail -f gibi sürekli izler.
    Yeni satırlarda tehdit tespiti de yapar.
    """
    def __init__(self, parent, **kwargs):
        super().__init__(parent, fg_color=COLORS['bg_panel'], **kwargs)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        self._monitor = None     # LogTailHandler örneği
        self._parser  = None     # Aktif parser
        self._analyzer = None   # BehavioralAnalyzer örneği
        self._line_count = 0    # İzlenen satır sayacı
        
        self._build()

    def _build(self):
        # ─ Kontrol çubuğu ────────────────────────────────────────────────
        ctrl = ctk.CTkFrame(self, fg_color=COLORS['bg_card'], corner_radius=8)
        ctrl.grid(row=0, column=0, sticky='ew', padx=10, pady=10)
        ctrl.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(ctrl, text="📡 GERÇEK ZAMANLI İZLEME",
                    font=("Consolas", 12, "bold"),
                    text_color=COLORS['accent']).grid(row=0, column=0, padx=10, pady=8)
        
        # Dosya yolu girişi
        self._path_entry = ctk.CTkEntry(ctrl,
                                       placeholder_text="İzlenecek log dosyasını seç...",
                                       font=("Consolas", 10),
                                       fg_color=COLORS['bg_input'],
                                       border_color=COLORS['border'],
                                       text_color=COLORS['text'])
        self._path_entry.grid(row=0, column=1, padx=5, pady=8, sticky='ew')
        
        ctk.CTkButton(ctrl, text="📁 Gözat",
                     font=("Consolas", 10),
                     fg_color=COLORS['bg_input'],
                     border_color=COLORS['border'], border_width=1,
                     hover_color=COLORS['border'],
                     command=self._browse).grid(row=0, column=2, padx=5)
        
        self._start_btn = ctk.CTkButton(ctrl, text="▶ BAŞLAT",
                                        font=("Consolas", 11, "bold"),
                                        fg_color=COLORS['success'],
                                        hover_color='#009955',
                                        command=self._toggle)
        self._start_btn.grid(row=0, column=3, padx=5)
        
        self._count_lbl = ctk.CTkLabel(ctrl, text="Satır: 0",
                                       font=("Consolas", 10),
                                       text_color=COLORS['accent2'])
        self._count_lbl.grid(row=0, column=4, padx=10)
        
        # ─ Canlı çıktı metni ─────────────────────────────────────────────
        self._output = ctk.CTkTextbox(self, font=("Consolas", 9),
                                     fg_color=COLORS['bg_input'],
                                     text_color=COLORS['text_mono'],
                                     wrap='word')
        self._output.grid(row=1, column=0, sticky='nsew', padx=10, pady=(0,10))
        self._output.configure(state='disabled')

    def _browse(self):
        """Dosya seçme diyalogu."""
        path = filedialog.askopenfilename(
            title="İzlenecek Log Dosyasını Seç",
            filetypes=[("Log Dosyaları", "*.log *.txt *.json *.gz"), ("Tümü", "*.*")]
        )
        if path:
            self._path_entry.delete(0, 'end')
            self._path_entry.insert(0, path)

    def _toggle(self):
        """İzlemeyi başlat veya durdur."""
        if self._monitor:
            self._stop()
        else:
            self._start()

    def _start(self):
        """Log izlemeyi başlat."""
        path = self._path_entry.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("Hata", "Geçerli bir log dosyası seçin.")
            return
        
        try:
            # Format tespiti ve parser başlatma
            sample = list(LogFileReader.read_lines(path, max_lines=10))
            fmt = FormatDetector.from_filename(os.path.basename(path)) or \
                  FormatDetector.detect(sample)
            self._parser   = get_parser(fmt)
            self._analyzer = BehavioralAnalyzer()
            
            def on_new_line(line: str):
                """Her yeni satır için callback."""
                self._line_count += 1
                ts = datetime.now().strftime('%H:%M:%S')
                
                # Satırı parse et
                entry = self._parser.parse_line(line, self._line_count)
                
                # Tehdit analizi
                alert_txt = ""
                if entry:
                    new_alerts = self._analyzer.analyze_entry(entry)
                    for a in new_alerts:
                        sev = a.severity.name if hasattr(a.severity, 'name') else str(a.severity)
                        alert_txt += f"\n  ⚡ [{sev}] {a.name} — {a.description}\n"
                
                # UI güncellemesi main thread'de olmalı
                self.after(0, self._append, f"[{ts}] {line[:120]}\n", alert_txt)
                self.after(0, self._count_lbl.configure,
                          {'text': f"Satır: {self._line_count:,}"})
            
            # Tail handler'ı başlat
            self._monitor = LogTailHandler(path, on_new_line)
            self._monitor.start()
            
            self._start_btn.configure(text="⏹ DURDUR",
                                     fg_color=COLORS['critical'],
                                     hover_color='#990000')
            
            self._append(f"[{datetime.now().strftime('%H:%M:%S')}] "
                        f"▶ İzleme başladı: {os.path.basename(path)}\n", "")
            
        except Exception as e:
            messagebox.showerror("Hata", f"İzleme başlatılamadı:\n{e}")

    def _stop(self):
        """İzlemeyi durdur."""
        if self._monitor:
            self._monitor.stop()
            self._monitor = None
        self._start_btn.configure(text="▶ BAŞLAT",
                                 fg_color=COLORS['success'],
                                 hover_color='#009955')
        self._append(f"[{datetime.now().strftime('%H:%M:%S')}] ⏹ İzleme durduruldu.\n", "")

    def _append(self, text: str, alert_text: str):
        """Çıktı kutusuna metin ekle ve sona kay."""
        self._output.configure(state='normal')
        self._output.insert('end', text)
        if alert_text:
            self._output.insert('end', alert_text)
        
        # Bellek yönetimi: 1000 satırdan fazla tutma
        content = self._output.get('1.0', 'end').split('\n')
        if len(content) > 1000:
            self._output.delete('1.0', f'{len(content)-900}.0')
        
        self._output.see('end')   # Sona otomatik kayar
        self._output.configure(state='disabled')


class ChartsPanel(ctk.CTkFrame):
    """
    Matplotlib ile 6 farklı güvenlik grafiği çizen panel.
    - Top Source IPs (yatay bar grafik)
    - HTTP Status Codes (çubuk grafik, renkli)
    - Event Timeline (çizgi + dolgu grafik)
    - HTTP Methods (pasta grafik)
    - Severity Distribution (çubuk grafik)
    - Alert Severity (pasta grafik)
    """
    def __init__(self, parent, **kwargs):
        super().__init__(parent, fg_color=COLORS['bg_panel'], **kwargs)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        # Yer tutucu metin (veri gelmeden önce)
        self._placeholder = ctk.CTkLabel(
            self,
            text="📊 Grafik için önce bir log dosyası parse edin\n\n"
                 "Mevcut grafikler:\n"
                 "• Top Kaynak IP'ler\n"
                 "• HTTP Durum Kodu Dağılımı\n"
                 "• Event Timeline (saatlik)\n"
                 "• HTTP Metod Dağılımı\n"
                 "• Severity Dağılımı\n"
                 "• Alert Severity Dağılımı",
            font=("Consolas", 12),
            text_color=COLORS['text_dim'],
            justify='center',
        )
        self._placeholder.place(relx=0.5, rely=0.5, anchor='center')

    def render(self, stats: dict, alerts: list):
        """6 grafiği oluştur ve tkinter'a göm."""
        try:
            import matplotlib
            matplotlib.use('Agg')  # GUI'siz backend (tkinter'a gömeceğiz)
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        except ImportError:
            self._placeholder.configure(
                text="matplotlib kurulu değil.\npip install matplotlib")
            return
        
        # Yer tutucuyu gizle
        self._placeholder.place_forget()
        
        # Eski canvas'ı temizle
        for w in self.winfo_children():
            if hasattr(w, 'get_tk_widget'):
                w.get_tk_widget().destroy()
            elif w != self._placeholder:
                w.destroy()
        
        # Dark tema uygula
        plt.style.use('dark_background')
        fig, axes = plt.subplots(2, 3, figsize=(14, 8))
        fig.patch.set_facecolor('#060d1a')
        
        # 6 farklı aksent rengi
        palette = ['#00c8ff','#00ffaa','#ff6b35','#ffd700','#ff4444','#cc88ff']
        
        # ── 1. Top Source IPs ──────────────────────────────────────────────
        ax = axes[0, 0]
        ax.set_facecolor('#0a1628')
        top_ips = stats.get('top_source_ips_list', [])[:10]
        if top_ips:
            ips, counts = zip(*top_ips)
            bars = ax.barh([ip[:20] for ip in ips], counts, color=palette[0])
            ax.set_title('Top Kaynak IP\'ler', color='#00c8ff', pad=8)
            ax.tick_params(colors='#5a7a9a', labelsize=7)
            # Değerleri bar'ların yanına yaz
            for bar, count in zip(bars, counts):
                ax.text(bar.get_width() + max(counts)*0.01,
                       bar.get_y() + bar.get_height()/2,
                       str(count), va='center', color='#d0e4f8', fontsize=7)
        self._style_ax(ax)
        
        # ── 2. HTTP Status Codes ───────────────────────────────────────────
        ax = axes[0, 1]
        ax.set_facecolor('#0a1628')
        sc_raw = dict(stats.get('status_codes', {}))
        if sc_raw:
            codes  = [str(k) for k in sorted(sc_raw.keys())]
            counts = [sc_raw[int(k)] for k in codes]
            # HTTP durum koduna göre renk: 2xx=yeşil, 3xx=sarı, 4xx=turuncu, 5xx=kırmızı
            colors_sc = []
            for code in codes:
                ci = int(code)
                if   ci >= 500: colors_sc.append('#ff4444')
                elif ci >= 400: colors_sc.append('#ff8c00')
                elif ci >= 300: colors_sc.append('#ffd700')
                else:           colors_sc.append('#00ffaa')
            ax.bar(codes, counts, color=colors_sc)
            ax.set_title('HTTP Durum Kodları', color='#00c8ff', pad=8)
            ax.tick_params(colors='#5a7a9a', labelsize=8, rotation=45)
        self._style_ax(ax)
        
        # ── 3. Event Timeline ──────────────────────────────────────────────
        ax = axes[0, 2]
        ax.set_facecolor('#0a1628')
        timeline = dict(stats.get('timeline', {}))
        if timeline:
            sorted_tl = sorted(timeline.items())
            vals = [v for _, v in sorted_tl]
            # Çizgi + dolgu efekti
            ax.plot(range(len(vals)), vals, color='#00c8ff', linewidth=1.5)
            ax.fill_between(range(len(vals)), vals, alpha=0.2, color='#00c8ff')
            ax.set_title('Event Timeline (Saatlik)', color='#00c8ff', pad=8)
            ax.set_xticks([])  # X ekseni etiketleri karışır, gizle
            ax.tick_params(colors='#5a7a9a', labelsize=7)
        self._style_ax(ax)
        
        # ── 4. HTTP Methods (Pasta) ────────────────────────────────────────
        ax = axes[1, 0]
        ax.set_facecolor('#0a1628')
        methods = dict(stats.get('methods', {}))
        if methods:
            ax.pie(methods.values(), labels=methods.keys(),
                  colors=palette[:len(methods)],
                  autopct='%1.1f%%',
                  textprops={'color': '#d0e4f8', 'fontsize': 8},
                  pctdistance=0.85)
            ax.set_title('HTTP Metod Dağılımı', color='#00c8ff', pad=8)
        
        # ── 5. Severity Distribution ───────────────────────────────────────
        ax = axes[1, 1]
        ax.set_facecolor('#0a1628')
        sev_dist = dict(stats.get('severity_dist', {}))
        if sev_dist:
            sev_color_map = {
                'critical':'#ff2222', 'error':'#ff4444', 'high':'#ff8c00',
                'warning':'#ffd700', 'medium':'#ffd700', 'low':'#00bfff',
                'notice':'#00ffaa', 'info':'#888888', 'debug':'#444444'
            }
            colors_sev = [sev_color_map.get(k.lower(), '#888') for k in sev_dist.keys()]
            ax.bar(list(sev_dist.keys()), list(sev_dist.values()), color=colors_sev)
            ax.set_title('Severity Dağılımı', color='#00c8ff', pad=8)
            ax.tick_params(colors='#5a7a9a', labelsize=8, rotation=30)
        self._style_ax(ax)
        
        # ── 6. Alert Severity (Pasta) ──────────────────────────────────────
        ax = axes[1, 2]
        ax.set_facecolor('#0a1628')
        if alerts:
            alert_sev = Counter(
                a.severity.name if hasattr(a.severity, 'name') else str(a.severity)
                for a in alerts
            )
            alert_colors = {
                'CRITICAL':'#ff2222','HIGH':'#ff8c00',
                'MEDIUM':'#ffd700','LOW':'#00bfff','INFO':'#888'
            }
            colors_al = [alert_colors.get(k, '#888') for k in alert_sev.keys()]
            ax.pie(alert_sev.values(), labels=alert_sev.keys(),
                  colors=colors_al,
                  autopct='%1.1f%%',
                  textprops={'color': '#d0e4f8', 'fontsize': 8})
            ax.set_title(f'Alert Dağılımı ({len(alerts)} toplam)',
                        color='#00c8ff', pad=8)
        else:
            ax.text(0.5, 0.5, 'Alert Yok', ha='center', va='center',
                   color='#5a7a9a', transform=ax.transAxes, fontsize=12)
            ax.set_title('Alert Dağılımı', color='#00c8ff', pad=8)
        
        fig.tight_layout(pad=2.0)
        
        # Matplotlib figürünü tkinter canvas'ına göm
        canvas = FigureCanvasTkAgg(fig, master=self)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    @staticmethod
    def _style_ax(ax):
        """Grafik ekseni için ortak dark tema stili."""
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_color('#1a3a6a')
        ax.spines['left'].set_color('#1a3a6a')


# ═══════════════════════════════════════════════════════════════════════════════
#  BÖLÜM 16: ANA UYGULAMA SINIFI
#  Tüm UI bileşenlerini bir araya getirir.
#  Menü, tab'lar, header, status bar burada tanımlanır.
#  İş mantığı (parse, analiz) arka plan thread'lerinde çalışır.
# ═══════════════════════════════════════════════════════════════════════════════
class CyberLogParserApp:
    """
    CyberLogParser Pro 2026 — Ana Uygulama Sınıfı
    
    Mimari:
    - UI: CustomTkinter (main thread)
    - Parse + Analiz: background thread (UI donmasın diye)
    - Callback: after() ile main thread'e geri dön
    
    Tab yapısı:
    1. Dashboard  → Özet istatistikler + Top listeler
    2. Log Kayıtları → Filtrelenebilir log tablosu
    3. Tehdit Uyarıları → Renk kodlu alert listesi
    4. IOC Intel → IOC çıkarımı ve pattern analizi
    5. Canlı İzleme → Gerçek zamanlı tail -f
    6. Grafikler → 6 matplotlib grafiği
    """
    
    def __init__(self):
        # Ana pencere
        self.root = ctk.CTk()
        self.root.title("🛡 CyberLogParser Pro 2026 — SOC Edition")
        self.root.geometry("1400x900")
        self.root.minsize(1100, 700)
        self.root.configure(fg_color=COLORS['bg_main'])
        
        # Uygulama durumu
        self._entries:        list = []        # Parse edilen log entry'leri
        self._alerts:         list = []        # Tespit edilen alert'ler
        self._stats:          dict = {}        # İstatistikler
        self._current_file:   str  = ""        # Açık dosya yolu
        self._is_parsing:     bool = False     # Parse devam ediyor mu?
        
        # UI oluştur
        self._build_ui()

    def _build_ui(self):
        """Tüm UI bileşenlerini oluşturur."""
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(1, weight=1)
        
        self._build_header()
        self._build_tabs()
        self._build_statusbar()

    def _build_header(self):
        """
        Üst başlık çubuğu.
        Logo, dosya açma, parse, export, ayarlar butonları burada.
        """
        hdr = ctk.CTkFrame(self.root, fg_color=COLORS['bg_panel'],
                          corner_radius=0, height=56)
        hdr.grid(row=0, column=0, sticky='ew')
        hdr.grid_columnconfigure(2, weight=1)  # Orta boşluğu esnet
        hdr.grid_propagate(False)
        
        # Logo
        ctk.CTkLabel(hdr, text="🛡 CyberLogParser",
                    font=("Consolas", 16, "bold"),
                    text_color=COLORS['accent']).grid(row=0, column=0, padx=15, pady=10)
        
        ctk.CTkLabel(hdr, text="PRO 2026 · SOC Edition",
                    font=("Consolas", 9),
                    text_color=COLORS['text_dim']).grid(row=0, column=1)
        
        # Buton stili (tekrar kullanım için)
        btn = dict(font=("Consolas", 10), fg_color=COLORS['bg_input'],
                  border_color=COLORS['border'], border_width=1,
                  hover_color=COLORS['border'], height=32)
        
        # Dosya açma butonu
        ctk.CTkButton(hdr, text="📂 Dosya Aç",
                     command=self._open_file, **btn).grid(
                     row=0, column=3, padx=5, pady=10)
        
        # Parse butonu (ana işlem butonu)
        self._parse_btn = ctk.CTkButton(
            hdr, text="⚡ Parse Et",
            font=("Consolas", 10, "bold"),
            fg_color=COLORS['accent'], text_color='#000',
            hover_color='#0099cc', height=32,
            command=self._start_parse)
        self._parse_btn.grid(row=0, column=4, padx=5, pady=10)
        
        # Dışa aktarma butonu
        ctk.CTkButton(hdr, text="📊 Dışa Aktar",
                     command=self._show_export_menu, **btn).grid(
                     row=0, column=5, padx=5, pady=10)
        
        # Temizle butonu
        ctk.CTkButton(hdr, text="🗑 Temizle",
                     command=self._clear, **btn).grid(
                     row=0, column=6, padx=5, pady=10)
        
        # Progress bar (parse işlemi sırasında doldu)
        self._progress = ctk.CTkProgressBar(hdr, width=150,
                                           fg_color=COLORS['bg_input'],
                                           progress_color=COLORS['accent'])
        self._progress.grid(row=0, column=7, padx=10)
        self._progress.set(0)
        
        # Yüklenen dosya adı etiketi
        self._file_lbl = ctk.CTkLabel(hdr, text="Dosya yüklenmedi",
                                     font=("Consolas", 9),
                                     text_color=COLORS['text_dim'])
        self._file_lbl.grid(row=0, column=8, padx=10)

    def _build_tabs(self):
        """
        6 sekmeli ana içerik alanı.
        Her sekme ayrı bir panel sınıfıdır.
        """
        self._tabs = ctk.CTkTabview(
            self.root,
            fg_color=COLORS['bg_panel'],
            segmented_button_fg_color=COLORS['bg_card'],
            segmented_button_selected_color=COLORS['accent'],
            segmented_button_selected_hover_color='#0099cc',
            segmented_button_unselected_color=COLORS['bg_card'],
            segmented_button_unselected_hover_color=COLORS['border'],
            text_color=COLORS['text'],
            text_color_disabled=COLORS['text_dim'],
        )
        self._tabs.grid(row=1, column=0, sticky='nsew', padx=8, pady=(4,0))
        
        # Sekmeleri oluştur
        tab_names = [
            "📊 Dashboard",
            "📋 Log Kayıtları",
            "🚨 Uyarılar",
            "🔍 IOC Intel",
            "📡 Canlı İzleme",
            "📈 Grafikler",
        ]
        for name in tab_names:
            self._tabs.add(name)
        
        # Her sekmeye panel ekle
        def tab(name):
            t = self._tabs.tab(name)
            t.grid_columnconfigure(0, weight=1)
            t.grid_rowconfigure(0, weight=1)
            return t
        
        # Dashboard sekmesi
        self._dashboard = DashboardPanel(tab("📊 Dashboard"))
        self._dashboard.grid(row=0, column=0, sticky='nsew')
        
        # Log kayıtları sekmesi
        self._log_view = LogViewPanel(tab("📋 Log Kayıtları"))
        self._log_view.grid(row=0, column=0, sticky='nsew')
        
        # Uyarılar sekmesi
        self._alerts_panel = AlertsPanel(tab("🚨 Uyarılar"))
        self._alerts_panel.grid(row=0, column=0, sticky='nsew')
        
        # IOC sekmesi
        self._ioc_panel = IOCPanel(tab("🔍 IOC Intel"))
        self._ioc_panel.grid(row=0, column=0, sticky='nsew')
        
        # Canlı izleme sekmesi
        self._live_panel = LiveMonitorPanel(tab("📡 Canlı İzleme"))
        self._live_panel.grid(row=0, column=0, sticky='nsew')
        
        # Grafikler sekmesi
        self._charts = ChartsPanel(tab("📈 Grafikler"))
        self._charts.grid(row=0, column=0, sticky='nsew')

    def _build_statusbar(self):
        """
        Alt durum çubuğu.
        Sol: durum mesajı | Sağ: canlı saat
        """
        sb = ctk.CTkFrame(self.root, fg_color=COLORS['bg_card'],
                         corner_radius=0, height=28)
        sb.grid(row=2, column=0, sticky='ew')
        sb.grid_columnconfigure(1, weight=1)
        sb.grid_propagate(False)
        
        self._status_lbl = ctk.CTkLabel(sb,
                                       text="Hazır  |  CyberLogParser Pro 2026",
                                       font=("Consolas", 9),
                                       text_color=COLORS['text_dim'])
        self._status_lbl.grid(row=0, column=0, padx=10)
        
        # Canlı saat (her saniye güncellenir)
        self._clock_lbl = ctk.CTkLabel(sb, text="",
                                       font=("Consolas", 9),
                                       text_color=COLORS['accent2'])
        self._clock_lbl.grid(row=0, column=2, padx=10)
        self._tick()  # Saat güncellemesini başlat

    def _tick(self):
        """Saati her saniye günceller. Kendini reschedule eder."""
        self._clock_lbl.configure(
            text=datetime.now().strftime('🕐 %Y-%m-%d  %H:%M:%S'))
        self.root.after(1000, self._tick)  # 1 saniye sonra tekrar çağır

    def _set_status(self, msg: str):
        """Alt durum çubuğu mesajını günceller."""
        self._status_lbl.configure(text=msg)

    # ─── Dosya İşlemleri ─────────────────────────────────────────────────────

    def _open_file(self):
        """Dosya seç diyalogu açar."""
        path = filedialog.askopenfilename(
            title="Log Dosyası Aç",
            filetypes=[
                ("Tüm Log Dosyaları", "*.log *.txt *.json *.gz *.bz2 *.zip *.csv"),
                ("Apache/Nginx Logs",  "*.log"),
                ("JSON Logs",          "*.json"),
                ("Sıkıştırılmış",      "*.gz *.bz2 *.zip"),
                ("CSV",                "*.csv"),
                ("Tümü",               "*.*"),
            ]
        )
        if path:
            self._current_file = path
            fname = os.path.basename(path)
            fsize = os.path.getsize(path) / 1024 / 1024  # MB
            self._file_lbl.configure(text=f"📄 {fname} ({fsize:.1f} MB)")
            self._set_status(f"Yüklendi: {fname}")

    def _start_parse(self):
        """
        Parse işlemini başlatır.
        Arka plan thread'i kullanılır → UI donmaz.
        """
        if not self._current_file:
            messagebox.showinfo("Bilgi", "Önce bir log dosyası açın (📂 Dosya Aç)")
            return
        if self._is_parsing:
            return  # Zaten parse ediliyor
        
        self._is_parsing = True
        self._parse_btn.configure(state='disabled', text="⏳ Parse ediliyor...")
        self._progress.set(0)
        self._set_status("Log dosyası parse ediliyor...")
        
        # Arka plan thread'i başlat
        threading.Thread(target=self._parse_worker, daemon=True).start()

    def _parse_worker(self):
        """
        Arka plan thread'inde çalışan parse fonksiyonu.
        
        Adımlar:
        1. Dosyayı parse et (LogEntry listesi oluştur)
        2. Tehdit analizi yap (ThreatAlert listesi oluştur)
        3. İstatistikleri hesapla
        4. UI'ı main thread'de güncelle (after() ile)
        """
        try:
            def on_progress(current, total):
                """Parse ilerlemesini progress bar'a yansıt."""
                if total > 0:
                    pct = min(current / max(total, 1), 0.85)
                    self.root.after(0, self._progress.set, pct)
            
            # ─ Adım 1: Parse ──────────────────────────────────────────────
            entries, fmt = auto_parse_file(
                self._current_file,
                progress_cb=on_progress
            )
            
            self.root.after(0, self._set_status,
                           f"{len(entries):,} kayıt parse edildi ({fmt.value}) — Tehdit analizi...")
            
            # ─ Adım 2: Tehdit Analizi ─────────────────────────────────────
            analyzer = BehavioralAnalyzer()
            raw_alerts = []
            for entry in entries:
                raw_alerts.extend(analyzer.analyze_entry(entry))
            
            # Aynı alert'leri birleştir (IP + alert_id kombinasyonu)
            seen = set()
            unique_alerts = []
            for a in raw_alerts:
                key = (a.alert_id, a.source_ip, a.username)
                if key not in seen:
                    seen.add(key)
                    unique_alerts.append(a)
            
            # ─ Adım 3: İstatistikler ──────────────────────────────────────
            stats = StatisticsEngine.compute(entries)
            
            # ─ Adım 4: UI Güncellemesi (main thread) ──────────────────────
            self.root.after(0, self._update_ui, entries, unique_alerts, stats, fmt)
            
        except Exception as e:
            import traceback
            self.root.after(0, messagebox.showerror, "Parse Hatası",
                           f"Dosya parse edilemedi:\n\n{type(e).__name__}: {e}\n\n{traceback.format_exc()[:500]}")
            self.root.after(0, self._reset_btn)

    def _update_ui(self, entries, alerts, stats, fmt):
        """
        Parse tamamlandıktan sonra tüm UI bileşenlerini günceller.
        Bu fonksiyon main thread'de çalışır (after() ile çağrılır).
        """
        self._entries = entries
        self._alerts  = alerts
        self._stats   = stats
        
        self._progress.set(1.0)  # %100 tamamlandı
        
        # Her paneli güncelle
        self._dashboard.update(stats, alerts)
        self._log_view.populate(entries)
        self._alerts_panel.populate(alerts)
        self._ioc_panel.populate(entries)
        self._charts.render(stats, alerts)
        
        # Kritik alert sayısını hesapla
        crit_count = sum(
            1 for a in alerts
            if hasattr(a.severity, 'name') and a.severity.name == 'CRITICAL'
        )
        
        # Durum çubuğunu güncelle
        self._set_status(
            f"✅ {len(entries):,} kayıt | "
            f"🚨 {len(alerts)} uyarı ({crit_count} kritik) | "
            f"Format: {fmt.value}"
        )
        
        self._reset_btn()
        
        # Alert varsa otomatik olarak uyarılar sekmesine geç
        if alerts:
            self._tabs.set("🚨 Uyarılar")
        else:
            self._tabs.set("📊 Dashboard")

    def _reset_btn(self):
        """Parse butonunu sıfırla."""
        self._is_parsing = False
        self._parse_btn.configure(state='normal', text="⚡ Parse Et")

    # ─── Dışa Aktarma ────────────────────────────────────────────────────────

    def _show_export_menu(self):
        """Format seçme penceresi."""
        if not self._entries:
            messagebox.showinfo("Bilgi", "Dışa aktarmak için önce parse yapın.")
            return
        
        # Modal pencere
        win = ctk.CTkToplevel(self.root)
        win.title("Dışa Aktarma Formatı")
        win.geometry("300x330")
        win.configure(fg_color=COLORS['bg_panel'])
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()
        
        ctk.CTkLabel(win, text="📊 DIŞA AKTARMA FORMATI",
                    font=("Consolas", 12, "bold"),
                    text_color=COLORS['accent']).pack(pady=15)
        
        # Her format için buton
        options = [
            ("🌐 HTML Rapor",     '.html', self._export_html),
            ("📑 PDF Rapor",      '.pdf',  self._export_pdf),
            ("📊 Excel (XLSX)",   '.xlsx', self._export_excel),
            ("📝 JSON (Tam)",     '.json', self._export_json),
            ("📋 CSV (Kayıtlar)", '.csv',  self._export_csv_entries),
            ("⚠ CSV (Uyarılar)", '.csv',  self._export_csv_alerts),
        ]
        
        for label, ext, fn in options:
            def _make(w, f, e):
                return lambda: (w.destroy(), self._do_export(f, e))
            ctk.CTkButton(win, text=label,
                         font=("Consolas", 11),
                         fg_color=COLORS['bg_card'],
                         border_color=COLORS['border'], border_width=1,
                         hover_color=COLORS['border'],
                         command=_make(win, fn, ext)).pack(
                         pady=3, padx=20, fill='x')

    def _do_export(self, fn, ext: str):
        """Dosya kaydetme diyalogu aç ve export fonksiyonunu çalıştır."""
        path = filedialog.asksaveasfilename(
            defaultextension=ext,
            filetypes=[(ext[1:].upper() + " Dosyası", f"*{ext}"), ("Tümü", "*.*")],
            initialfile=f"cyberlogparser_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        )
        if not path:
            return
        
        try:
            gen = ReportGenerator(self._entries, self._alerts, self._stats)
            fn(gen, path)
            messagebox.showinfo("Tamamlandı", f"✅ Dosya kaydedildi:\n{path}")
            self._set_status(f"Dışa aktarıldı: {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Hata", f"Dışa aktarma başarısız:\n{e}")

    # Export yardımcı fonksiyonları (her format için ayrı)
    def _export_html(self, gen: ReportGenerator, path: str):
        gen.export_html(path)

    def _export_pdf(self, gen: ReportGenerator, path: str):
        gen.export_pdf(path)

    def _export_excel(self, gen: ReportGenerator, path: str):
        gen.export_excel(path)

    def _export_json(self, gen: ReportGenerator, path: str):
        gen.export_json(path)

    def _export_csv_entries(self, gen: ReportGenerator, path: str):
        gen.export_csv(path)

    def _export_csv_alerts(self, gen: ReportGenerator, path: str):
        gen.export_alerts_csv(path)

    # ─── Diğer Aksiyonlar ────────────────────────────────────────────────────

    def _clear(self):
        """Tüm verileri temizle."""
        if self._entries and messagebox.askyesno("Temizle", "Tüm veriler silinsin mi?"):
            self._entries = []
            self._alerts  = []
            self._stats   = {}
            self._log_view.populate([])
            self._alerts_panel.populate([])
            self._dashboard.update({}, [])
            self._progress.set(0)
            self._file_lbl.configure(text="Dosya yüklenmedi")
            self._set_status("Temizlendi")

    def run(self):
        """Uygulamayı başlat."""
        self.root.mainloop()


# ═══════════════════════════════════════════════════════════════════════════════
#  GİRİŞ NOKTASI
#  Bu dosya direkt çalıştırıldığında burası çalışır.
#  python CyberLogParser.py
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("   CyberLogParser Pro 2026 — SOC Edition")
    print("   Advanced Log Analysis & Threat Detection")
    print("=" * 60)
    print()
    
    # Uygulamayı başlat
    app = CyberLogParserApp()
    app.run()